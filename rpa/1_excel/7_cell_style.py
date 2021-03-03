from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("score_2.xlsx")
ws = wb.active

# 번호, 컴퓨터 수학

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 컴퓨터
c1 = ws["C1"] # 수학

# a열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color = "FF0000", italic = True, bold = True) # 글자 색은 빨간색, 기울임, 굵게 적용
b1.font = Font(color = "CC33FF", name = "Arial", italic = True, bold = True, strike= True) # 폰트 변경, 취소선 적용
c1.font = Font(color = "0000FF", size= 20 , italic = True, bold = True, underline= "single") # 글자 크기를 20으로, 밑줄을 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right = Side(style="thin"), top = Side(style="thin"), bottom = Side(style="thin"))

a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal= "center", vertical= "center") # 데이터 서식 정렬
        # center, left, right, top, bottom
        if cell.column == 1: # A번호열은 제외
            continue
        # cell이 정수형 데이터고 90점 보다 높으면
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor="00FF00", fill_type = "solid") # 배경을 초록색으로
            cell.font = Font(color="FF0000")

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정
wb.save("sample_style.xlsx")