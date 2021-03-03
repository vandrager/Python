from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

'''
# 수식 그대로 가져오고 있음
for row in ws.values:
    for cell in row:
        print(cell)
'''
# 수식이 아닌 실제 데이터를 가져옴
# evaluate 되지 않은 상태의 데이터는 none이라고 함
# 따라서, 데이터를 가져올 때 엑셀 파일을 열었다가 다시 저장하고 불러와야함
wb = load_workbook("sample_formula.xlsx", data_only= True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 셀 병합하기
ws.merge_cells("B2:D2")
ws['B2'].value = "Merged Cell"
ws.unmerge_cells("B2:D2") # 셀 병합 해제
wb.save("sample_merge.xlsx")