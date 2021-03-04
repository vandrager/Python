from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd
# wb = Workbook()
# ws = wb.active

# img = Image("img.png")

# ws.add_image(img, "C3")

# # import error you must install Pillow to fetch image...
# # pip install pillow
# wb.save("sample_image.xlsx")

# quiz
# from openpyxl import load_workbook
# wb = load_workbook("data.xlsx")
# ws = wb.active
'''
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores: # 기존 성적 데이터 넣기
    ws.append(s)

# 1. 퀴즈2 점수를 10 으로 수정
for idx, cell in enumerate(ws["D"]):
    if idx != 0:
        cell.value = 10


# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

# a b c d e f g h i 
# 1 2 3 4 5 6 7 8 9
for idx, score in enumerate(scores, start=2):
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # SUM(B2:G2)
    # SUM(B3:G3)...
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    # - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
    grade = None # 성적 
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    # 3. 출석이 5 미만인 학생은 총점 상관없이 F 
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade # I 열에 성적 정보

wb.save("scores.xlsx")
'''

# wb.save("scores.xlsx")

# 판다스로 함 해볼까
'''
df = pd.read_excel("data.xlsx")
df['총점'] = 0
df['성적'] = 0
df.columns = ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "성적"]
df['퀴즈2'] = 10
for i in range(len(df['총점'])):
    sum_value = df['출석'][i] + df['퀴즈1'][i] + df['퀴즈2'][i] + df['중간고사'][i] + df['기말고사'][i] + df['프로젝트'][i]
    df['총점'][i] = sum_value
    if df['출석'][i] < 5:
        df['성적'][i] = "F"
    elif sum_value >= 90:
        df['성적'][i] = "A"
    elif sum_value >= 80:
        df['성적'][i] = "B"
    elif sum_value >= 70:
        df['성적'][i] = "C"
    else:
        df['성적'][i] = "D"
# 1. 퀴즈2 점수를 10 으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F
df.to_excel("scores.xlsx")
'''

from openpyxl import load_workbook
wb = load_workbook("scores.xlsx", data_only= True)
ws = wb.active
# ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
print(ws.cell(row = 3, column = 4).value)
for x in range(1, 12):
    for y in range(1, 10):
        # ws.cell(row = x, column = y, value = randint(0, 100)) # 0 ~ 100 사이의 랜덤 숫자 삽입
        print(ws.cell(row = x, column = y).value, end = " ")
    if ws.cell(row = x, column = 9).value == "A":
        print("<- 만나면 좋은 친구~")
    print()