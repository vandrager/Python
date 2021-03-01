from openpyxl import load_workbook

wb = load_workbook("score_2.xlsx")
ws = wb.active

# for row in ws.iter_rows(min_row = 2):
#     if int(row[1].value) > 80:
#         print(row[0].value, "번 학생은 영어 천재..!", row[1].value, "점")

# for row in ws.iter_rows(max_row = 1):
#     for cell in row:
#         if cell.value == "영어":
#             cell.value = "컴퓨터"

# wb.save("score_2.xlsx")

# 엑셀 행, 열 삽입
ws.insert_rows(8) # 8번째 줄이 비워짐
# ws.insert_rows(8, 5) # 8번째 줄부터 5개를 비움
ws.insert_cols(2) # 2번째 열이 비워짐 새로운 빈 열을 추가
# ws.insert_cols(2, 3) # 2번째 열이 비워짐 새로운 빈 열을 3개 추가


# 엑셀 행, 열 삭제
ws.delete_rows(8)
ws.delete_cols(2)
# ws.delete_cols(2, 2) # 여러개씩 삭제도 가능

# 엑셀 행, 열 이동

wb.save("score_2.xlsx")