from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string # 좌표정보를 가져오고 싶다면 사용
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_b = ws["B"] # 영어 컬럼만 가지고 오기
# print(col_b)
# for c in col_b:
#     print(c.value)

# col_range = ws["B:C"] # 영어, 수학 컬럼 함께 가지고 오기
# for r in col_range:
#     for c in r:
#         print(c.value)

# row_title = ws[1] # 1번째 로우만 가지고 오기

# for r in row_title:
#     print(r.value)

# row_range = ws[2:6] # 1번째 로우는 제외하고 2번째에서 6번째 줄까지 가지고 옴

# for r in row_range:
#     for c in r:
#         print(c.value, end = " ")
#     print()

row_range = ws[2:ws.max_row] # 1번째 로우는 제외하고 마지막 줄까지 가지고 옴

for r in row_range:
    for c in r:
        # print(c.coordinate, end = " ")
        xy = coordinate_from_string(c.coordinate)
        # print(xy, end = " ")
        print(xy[0], end = " ") # A
        print(xy[1], end = " ") # 2
    print()

# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value)

# 전체 columns
print(tuple(ws.columns))
for col in tuple(ws.columns): #컬럼명만 가져와보자
    print(col[0].value)

for row in ws.iter_rows(): # 전체 row
    print(row)
'''
(<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
(<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>)
(<Cell 'Sheet'.A3>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.C3>)
(<Cell 'Sheet'.A4>, <Cell 'Sheet'.B4>, <Cell 'Sheet'.C4>)
(<Cell 'Sheet'.A5>, <Cell 'Sheet'.B5>, <Cell 'Sheet'.C5>)
(<Cell 'Sheet'.A6>, <Cell 'Sheet'.B6>, <Cell 'Sheet'.C6>)
(<Cell 'Sheet'.A7>, <Cell 'Sheet'.B7>, <Cell 'Sheet'.C7>)
(<Cell 'Sheet'.A8>, <Cell 'Sheet'.B8>, <Cell 'Sheet'.C8>)
(<Cell 'Sheet'.A9>, <Cell 'Sheet'.B9>, <Cell 'Sheet'.C9>)
(<Cell 'Sheet'.A10>, <Cell 'Sheet'.B10>, <Cell 'Sheet'.C10>)
(<Cell 'Sheet'.A11>, <Cell 'Sheet'.B11>, <Cell 'Sheet'.C11>)
'''
for row in ws.iter_rows(): # 수학 row 값
    print(row[2].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지 # 좌우 좌우로 데이터 가져옴
for row in ws.iter_rows(min_row = 1, max_row = 5, min_col = 2, max_col = 3): # 전체 row 슬라이싱
    print(row[0].value, row[1].value) # 수학, 영어
    print(row)

# 상하 상하로 데이터 가져옴
for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 2, max_col = 3): # 전체 column 슬라이싱
    print(col[0].value, col[1].value) # 수학, 영어
    print(col)
wb.save("score.xlsx")


