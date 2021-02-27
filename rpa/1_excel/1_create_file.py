from openpyxl import Workbook

wb = Workbook() # 새 워크북을 생성
ws = wb.active # 현재 활성화된 sheet를 가져옴
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "ybSheet" # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경
ws1 = wb.create_sheet("srSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("newSheet", 2) # 2번째 인덱스에 시트 생성
new_ws = wb['newSheet'] # dict 형태로 sheet에 접근
print(wb.sheetnames) # 모든 시트 이름 확인

# 시트 복사
new_ws['A1'] = 'TEST'
target = wb.copy_worksheet(new_ws)
target.title = "copied sheet"
ws.append(["bye", "bye"])
wb.save("sample.xlsx") # 워크북 저장

