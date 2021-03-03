from openpyxl import load_workbook

wb = load_workbook("score_2.xlsx")
ws = wb.active

# 번호, 컴퓨터, 수학
# 번호, (국어), 컴퓨터, 수학

# ws.move_range("B1:C11", rows=0, cols=1) # 열만 오른쪽으로 한 칸 이동
# ws["B1"].value = "국어"
# ws.move_range("B1:C11", rows=0, cols=-1) # 열만 왼쪽으로 다시 한 칸 이동
# wb.save("score_2.xlsx")

from openpyxl.chart import BarChart, Reference, LineChart 
# # B2:C11까지의 데이터를 엑셀 차트로 생성
# bar_value = Reference(ws, min_row = 2, max_row = 11, min_col = 2, max_col = 3)
# bar_chart = BarChart() # 차트 종류 설정
# bar_chart.add_data(bar_value) # 차트 데이터 추가
# ws.add_chart(bar_chart, "E1") # 차트 넣을 위치 정의

# B1:C11까지의 데이터를 엑셀 햐차트로 생성
line_value = Reference(ws, min_row = 1, max_row = 11, min_col = 2, max_col = 3)
line_chart = LineChart() # 차트 종류 설정
line_chart.add_data(line_value, titles_from_data=True) # 차트 데이터 추가 # 계열 > 영어, 수학(제목에서 가져옴)
line_chart.title = "성적표" # 차트 제목
line_chart.style = 20 # 미리 정의된 스타일을 적용, 사용자 개별 지정도 가능
line_chart.y_axis.title = "점수"
line_chart.x_axis.title = "번호"
ws.add_chart(line_chart, "E15") # 차트 넣을 위치 정의

wb.save("score_2.xlsx")

# 차트 스타일/내용에 대한 자세한 매뉴얼은 아래 링크 참고
# https://openpyxl.readthedocs.io/en/stable/charts/introduction.html