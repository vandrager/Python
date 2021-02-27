from openpyxl import Workbook
from random import *
import os
wb = Workbook()
ws = wb.active
ws.title = "ybsheet"

# A1 셀에 1이라는 값을 입력
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['A4'] = 4
ws['A5'] = 5
ws['A6'] = 6

print(ws['A1'])
print(ws['A1'].value) # 값을 출력
print(ws['A10'].value) # 없으면 none 출력

# row = 1, 2, 3
# column = A(1), B(2), C(3)
print(ws.cell(row = 1, column = 1).value) # ws['A1'].value
print(ws.cell(row = 1, column = 2).value) # ws['B1'].value

c = ws.cell(column = 3, row = 1, value = 10) # ws['c1'].value = 10과 동일
print(c.value)

# 반복문을 사용해 숫자 채우기 
index = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row = x, column = y, value = randint(0, 100)) # 0 ~ 100 사이의 랜덤 숫자 삽입
        ws.cell(row = x, column = y, value = index)
        index += 1
wb.save("sample2.xlsx")